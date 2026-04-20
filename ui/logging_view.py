"""Data Logger tab — industrial SCADA-grade UI.

Features:
- Status chip bar (Running / Stopped / Stale / Disk Warning / Error)
- One-click Start / Stop / Pause / Resume / Save Settings
- Key presets: Essential, Power Quality, Energy Billing, Full Diagnostics
- Per-source parameter selection with search/filter + Select All/None
- Shows configured interval, last-write timestamp, failed writes, skipped samples
- Validates interval bounds and storage path
- Retention control + cleanup action
- Open log folder / copy path quick actions
- Recent-record preview (last 5 rows)
- Toast-style feedback for operator actions
"""

from __future__ import annotations

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime, timedelta
import os
import csv
import subprocess

from core.logging_engine import LoggingEngine, canonicalize_log_key
from ui.styles import get_theme
from ui.key_registry import (
    COMMON_KEYS,
    HARMONIC_LOG_KEYS,
    HARMONIC_SUMMARY_KEYS,
    canonical_key,
    key_label,
)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

ALL_LOG_KEYS = list(dict.fromkeys(COMMON_KEYS + HARMONIC_SUMMARY_KEYS + HARMONIC_LOG_KEYS + [
    "Total_kW", "Total_kVA", "Total_kVAr", "PF_total", "RunHour_total", "Today_kWh"
]))

DEFAULT_LOG_KEYS = ["kW", "kVA", "kVAr", "Vavg", "Frequency", "I1", "I2", "I3"]

# Key presets (name → list of canonical keys)
KEY_PRESETS = {
    "Essential": ["kW", "kVA", "kVAr", "Vavg", "Iavg", "PFavg", "Frequency"],
    "Power Quality": ["V1", "V2", "V3", "I1", "I2", "I3", "PF1", "PF2", "PF3",
                      "THD_V1", "THD_V2", "THD_V3", "THD_I1", "THD_I2", "THD_I3",
                      "Vavg", "Iavg", "PFavg", "Frequency"],
    "Energy Billing": ["Import_kWh", "Export_kWh", "kWh", "kVAh",
                       "Today_kWh", "kW", "kVA", "PFavg"],
    "Individual Harmonics": HARMONIC_SUMMARY_KEYS + HARMONIC_LOG_KEYS,
    "Full Diagnostics": list(dict.fromkeys(DEFAULT_LOG_KEYS + ALL_LOG_KEYS)),
}

ONE_SHOT_PRESETS = [
    ("5 min", 5),
    ("15 min", 15),
    ("30 min", 30),
    ("1 hour", 60),
    ("2 hours", 120),
    ("4 hours", 240),
    ("12 hours", 720),
    ("1 day", 1440),
    ("2 days", 2880),
    ("5 days", 7200),
    ("7 days", 10080),
]

INTERVALS = [
    ("5 sec", 5),
    ("10 sec", 10),
    ("30 sec", 30),
    ("1 min", 60),
    ("3 min", 180),
    ("5 min", 300),
    ("10 min", 600),
    ("15 min", 900),
    ("30 min", 1800),
    ("1 hour", 3600),
]

MODES = [
    ("Instant", "instant"),
    ("Minimum", "min"),
    ("Average", "avg"),
    ("Maximum", "max"),
]

# Chip colours (dark theme)
_CHIP = {
    "RUNNING":       ("#064e3b", "#34d399"),   # bg, fg
    "STOPPED":       ("#1f1f23", "#6b7280"),
    "SCHEDULED_OFF": ("#422006", "#fb923c"),
    "PAUSED":        ("#1e1b4b", "#a5b4fc"),
    "ERROR":         ("#450a0a", "#f87171"),
    "STALE":         ("#1c1917", "#78716c"),
    "DISK_WARN":     ("#451a03", "#fbbf24"),
    "DISABLED":      ("#1f1f23", "#6b7280"),
}
_MUTED  = "#9a9490"
_ACCENT = "#4da6ff"


def _chip_text(state: str) -> str:
    return {
        "RUNNING":       "● RUNNING",
        "STOPPED":       "○ STOPPED",
        "SCHEDULED_OFF": "◷ SCHED OFF",
        "PAUSED":        "⏸ PAUSED",
        "ERROR":         "✕ ERROR",
        "STALE":         "~ STALE",
        "DISK_WARN":     "⚠ DISK",
        "DISABLED":      "— DISABLED",
    }.get(state, state)


def _fmt_age(dt: datetime | None) -> str:
    if dt is None:
        return "—"
    secs = (datetime.now() - dt).total_seconds()
    if secs < 60:
        return f"{int(secs)}s ago"
    if secs < 3600:
        return f"{int(secs // 60)}m ago"
    return dt.strftime("%H:%M:%S")


# ---------------------------------------------------------------------------
# LoggingTab
# ---------------------------------------------------------------------------

class LoggingTab(ttk.Frame):
    """Industrial SCADA logging control panel."""

    def __init__(self, parent, logging_engine: LoggingEngine, cfg: dict, snapshot_bus=None):
        super().__init__(parent, padding=0)
        self.logging_engine = logging_engine
        self.cfg = cfg if cfg is not None else {}
        self.columnconfigure(0, weight=1)
        self.rowconfigure(3, weight=1)

        self._dirty = False
        self._source_rows: dict[str, dict] = {}
        self.schedule_mode_var = tk.StringVar(value="always")
        self._last_total_keys: tuple[str, ...] = ()
        self._last_meter_ids: tuple[int, ...] = ()
        self._paused = False
        self._toast_job: str | None = None

        self._build_header()
        self._build_control_bar()
        self._build_status_panel()
        self._build_main_body()
        self._build_schedule_panel()

        self.logging_engine.configure(cfg, snapshot_bus=snapshot_bus)
        self._load_from_cfg()

    # =========================================================================
    # UI builders
    # =========================================================================

    def _build_header(self):
        t = get_theme()
        wrap = tk.Frame(self, bg=t.bg)
        wrap.grid(row=0, column=0, sticky="ew", padx=12, pady=(12, 8))
        card = tk.Frame(wrap, bg="#0b1220", highlightthickness=1, highlightbackground=t.border)
        card.pack(fill="x")
        tk.Frame(card, bg=t.warn, height=3).pack(fill="x", side="top")
        body = tk.Frame(card, bg="#0b1220")
        body.pack(fill="x", padx=14, pady=12)
        tk.Label(body, text="Data Logger", bg="#0b1220", fg=t.text, font=("Segoe UI", 17, "bold")).pack(side="left")
        tk.Label(
            body,
            text="Historian setup, source selection, schedules and recent record preview.",
            bg="#0b1220",
            fg=t.text_muted,
            font=("Segoe UI", 9),
        ).pack(side="left", padx=(12, 0), pady=(5, 0))

    def _build_control_bar(self):
        bar = ttk.Frame(self, padding=(10, 8, 10, 6))
        bar.grid(row=1, column=0, sticky="ew")

        # Left: action buttons
        btn_frm = ttk.Frame(bar)
        btn_frm.pack(side="left")

        self.btn_start = ttk.Button(
            btn_frm, text="▶  Start", style="Accent.TButton",
            command=self._on_start, width=10,
        )
        self.btn_start.pack(side="left", padx=(0, 4))

        self.btn_stop = ttk.Button(
            btn_frm, text="■  Stop", command=self._on_stop, width=9,
        )
        self.btn_stop.pack(side="left", padx=4)

        self.btn_pause = ttk.Button(
            btn_frm, text="⏸  Pause", command=self._on_pause, width=9,
        )
        self.btn_pause.pack(side="left", padx=4)

        ttk.Separator(bar, orient="vertical").pack(side="left", fill="y", padx=10)

        # Interval + mode
        ttk.Label(bar, text="Interval:").pack(side="left", padx=(0, 3))
        self.cmb_interval = ttk.Combobox(
            bar, values=[lbl for lbl, _ in INTERVALS],
            state="readonly", width=9,
        )
        self.cmb_interval.set("10 sec")
        self.cmb_interval.pack(side="left", padx=(0, 8))
        self.cmb_interval.bind("<<ComboboxSelected>>", self._on_interval_changed)

        ttk.Label(bar, text="Mode:").pack(side="left", padx=(0, 3))
        self.cmb_mode = ttk.Combobox(
            bar, values=[lbl for lbl, _ in MODES],
            state="readonly", width=9,
        )
        self.cmb_mode.set("Instant")
        self.cmb_mode.pack(side="left", padx=(0, 8))
        self.cmb_mode.bind("<<ComboboxSelected>>", self._on_mode_changed)

        self.keep_awake_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(
            bar, text="Keep PC Awake",
            variable=self.keep_awake_var,
            command=self._on_keep_awake_changed,
        ).pack(side="left", padx=(0, 8))

        # Right: save + toast
        right = ttk.Frame(bar)
        right.pack(side="right")

        self.lbl_toast = ttk.Label(right, text="", foreground="#34d399",
                                   font=("Segoe UI", 9))
        self.lbl_toast.pack(side="right", padx=(8, 0))

        self.btn_save = ttk.Button(
            right, text="💾  Save Settings",
            command=self._on_apply_settings, width=14,
        )
        self.btn_save.pack(side="right")
        self.btn_save.state(["disabled"])

        self.lbl_dirty_dot = ttk.Label(right, text="●", foreground="#1f1f23",
                                       font=("Segoe UI", 12))
        self.lbl_dirty_dot.pack(side="right", padx=(0, 4))

        ttk.Separator(self, orient="horizontal").grid(
            row=1, column=0, sticky="ews", padx=0)

    def _build_status_panel(self):
        pnl = ttk.Frame(self, padding=(10, 8, 10, 6))
        pnl.grid(row=2, column=0, sticky="ew")
        pnl.columnconfigure(1, weight=1)

        # ── Chip row ────────────────────────────────────────────────────────
        chips = ttk.Frame(pnl)
        chips.grid(row=0, column=0, columnspan=3, sticky="w", pady=(0, 6))

        self.chip_state = tk.Label(
            chips, text="○ STOPPED", padx=8, pady=2,
            bg=_CHIP["STOPPED"][0], fg=_CHIP["STOPPED"][1],
            font=("Segoe UI", 9, "bold"), relief="flat", bd=0,
        )
        self.chip_state.pack(side="left", padx=(0, 4))

        self.chip_disk = tk.Label(
            chips, text="", padx=6, pady=2,
            bg=_CHIP["DISK_WARN"][0], fg=_CHIP["DISK_WARN"][1],
            font=("Segoe UI", 9, "bold"), relief="flat", bd=0,
        )
        # chip_disk packed on demand

        self.chip_err = tk.Label(
            chips, text="", padx=6, pady=2,
            bg=_CHIP["ERROR"][0], fg=_CHIP["ERROR"][1],
            font=("Segoe UI", 9, "bold"), relief="flat", bd=0,
        )

        # ── Metrics grid ────────────────────────────────────────────────────
        metrics = ttk.Frame(pnl)
        metrics.grid(row=1, column=0, sticky="w")

        def _stat(parent, label, row, col):
            ttk.Label(parent, text=label, foreground=_MUTED,
                      font=("Segoe UI", 8)).grid(
                row=row, column=col * 2, sticky="w", padx=(0, 4), pady=1)
            var = tk.StringVar(value="—")
            lbl = ttk.Label(parent, textvariable=var,
                            font=("Consolas", 9))
            lbl.grid(row=row, column=col * 2 + 1, sticky="w",
                     padx=(0, 20), pady=1)
            return var

        self._v_next    = _stat(metrics, "Next write:",  0, 0)
        self._v_last    = _stat(metrics, "Last write:",  0, 1)
        self._v_rows    = _stat(metrics, "Rows today:",  0, 2)
        self._v_errs    = _stat(metrics, "Write errors:", 1, 0)
        self._v_skipped = _stat(metrics, "Dropped:",     1, 1)
        self._v_storage = _stat(metrics, "Storage:",     1, 2)

        # ── File row ────────────────────────────────────────────────────────
        file_row = ttk.Frame(pnl)
        file_row.grid(row=2, column=0, columnspan=3, sticky="ew", pady=(4, 0))
        pnl.columnconfigure(0, weight=1)

        ttk.Label(file_row, text="File:", foreground=_MUTED,
                  font=("Segoe UI", 8)).pack(side="left")
        self.lbl_file = ttk.Label(file_row, text="No file",
                                  font=("Consolas", 8), foreground="#94a3b8")
        self.lbl_file.pack(side="left", padx=(4, 0), fill="x", expand=True)

        ttk.Button(file_row, text="Copy Path",
                   command=self._copy_path, width=9).pack(side="right", padx=(4, 0))
        ttk.Button(file_row, text="Open Folder",
                   command=self._open_folder, width=10).pack(side="right", padx=(4, 0))
        ttk.Button(file_row, text="Export Excel…",
                   command=self._export_excel, width=12).pack(side="right", padx=(4, 0))

        ttk.Separator(self, orient="horizontal").grid(
            row=2, column=0, sticky="ews")

    def _build_main_body(self):
        """Paned body: left = sources+presets, right = recent records."""
        paned = ttk.PanedWindow(self, orient="horizontal")
        paned.grid(row=3, column=0, sticky="nsew", padx=0, pady=0)

        # ── Left pane: sources ───────────────────────────────────────────────
        left = ttk.Frame(paned, padding=(10, 8, 6, 8))
        paned.add(left, weight=3)
        left.columnconfigure(0, weight=1)
        left.rowconfigure(1, weight=1)

        # Presets + search bar
        top = ttk.Frame(left)
        top.grid(row=0, column=0, sticky="ew", pady=(0, 6))

        ttk.Label(top, text="Presets:", foreground=_MUTED,
                  font=("Segoe UI", 8)).pack(side="left", padx=(0, 4))
        for name in KEY_PRESETS:
            ttk.Button(
                top, text=name,
                command=lambda n=name: self._apply_preset(n),
                width=len(name) + 2,
            ).pack(side="left", padx=2)

        ttk.Separator(top, orient="vertical").pack(side="left", fill="y", padx=8)

        ttk.Label(top, text="🔍", font=("Segoe UI", 9)).pack(side="left")
        self._search_var = tk.StringVar()
        self._search_var.trace_add("write", lambda *_: self._on_search_changed())
        ent_search = ttk.Entry(top, textvariable=self._search_var, width=16)
        ent_search.pack(side="left", padx=(2, 4))

        self.btn_select_all = ttk.Button(
            top, text="All", command=self._select_all_visible, width=4)
        self.btn_select_all.pack(side="left", padx=2)
        ttk.Button(
            top, text="None", command=self._select_none_visible, width=5
        ).pack(side="left", padx=2)

        # Scrollable sources
        sources_frm = ttk.LabelFrame(left, text="Logging Sources")
        sources_frm.grid(row=1, column=0, sticky="nsew")
        sources_frm.columnconfigure(0, weight=1)
        sources_frm.rowconfigure(0, weight=1)

        canvas = tk.Canvas(sources_frm, highlightthickness=0)
        try:
            bg = ttk.Style().lookup("TFrame", "background") or "#0b1220"
            canvas.configure(bg=bg)
        except Exception:
            pass
        canvas.grid(row=0, column=0, sticky="nsew")
        sb = ttk.Scrollbar(sources_frm, orient="vertical", command=canvas.yview)
        sb.grid(row=0, column=1, sticky="ns")
        canvas.configure(yscrollcommand=sb.set)

        self._source_container = ttk.Frame(canvas)
        canvas.create_window((0, 0), window=self._source_container, anchor="nw")
        self._source_container.bind(
            "<Configure>",
            lambda e: canvas.configure(scrollregion=canvas.bbox("all")),
        )

        # ── Right pane: recent records ───────────────────────────────────────
        right = ttk.Frame(paned, padding=(6, 8, 10, 8))
        paned.add(right, weight=2)
        right.columnconfigure(0, weight=1)
        right.rowconfigure(1, weight=1)

        hdr = ttk.Frame(right)
        hdr.grid(row=0, column=0, sticky="ew", pady=(0, 4))
        ttk.Label(hdr, text="Recent Records",
                  font=("Segoe UI", 9, "bold")).pack(side="left")
        self.lbl_preview_info = ttk.Label(hdr, text="", foreground=_MUTED,
                                          font=("Segoe UI", 8))
        self.lbl_preview_info.pack(side="left", padx=8)
        ttk.Button(hdr, text="Refresh", command=self._refresh_preview,
                   width=8).pack(side="right")

        frm_tree = ttk.Frame(right)
        frm_tree.grid(row=1, column=0, sticky="nsew")
        frm_tree.columnconfigure(0, weight=1)
        frm_tree.rowconfigure(0, weight=1)

        self._preview_tree = ttk.Treeview(
            frm_tree, show="headings", height=8,
            selectmode="none",
        )
        self._preview_tree.grid(row=0, column=0, sticky="nsew")
        sb_h = ttk.Scrollbar(frm_tree, orient="horizontal",
                              command=self._preview_tree.xview)
        sb_h.grid(row=1, column=0, sticky="ew")
        sb_v = ttk.Scrollbar(frm_tree, orient="vertical",
                              command=self._preview_tree.yview)
        sb_v.grid(row=0, column=1, sticky="ns")
        self._preview_tree.configure(
            xscrollcommand=sb_h.set, yscrollcommand=sb_v.set)
        self._preview_cols: tuple[str, ...] = ()

    def _build_schedule_panel(self):
        frm = ttk.LabelFrame(self, text="Schedule", padding=(10, 6))
        frm.grid(row=4, column=0, sticky="ew", padx=0, pady=(0, 0))

        row1 = ttk.Frame(frm)
        row1.pack(fill="x", pady=(0, 4))

        for text, val in [("Always ON", "always"),
                           ("Date/Time Range", "date_range"),
                           ("One-Shot", "one_shot")]:
            ttk.Radiobutton(
                row1, text=text,
                variable=self.schedule_mode_var, value=val,
                command=self._on_schedule_changed,
            ).pack(side="left", padx=(0, 12))

        self.time_frm = ttk.Frame(frm)
        self.time_frm.pack(fill="x", pady=(0, 4))

        ttk.Label(self.time_frm, text="From:").pack(side="left")
        self.ent_start_date = ttk.Entry(self.time_frm, width=11)
        self.ent_start_date.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.ent_start_date.pack(side="left", padx=2)
        self.ent_start_time = ttk.Entry(self.time_frm, width=6)
        self.ent_start_time.insert(0, "00:00")
        self.ent_start_time.pack(side="left", padx=2)

        ttk.Label(self.time_frm, text="To:", font=("Segoe UI", 9)).pack(
            side="left", padx=(10, 0))
        self.ent_end_date = ttk.Entry(self.time_frm, width=11)
        self.ent_end_date.insert(0, datetime.now().strftime("%Y-%m-%d"))
        self.ent_end_date.pack(side="left", padx=2)
        self.ent_end_time = ttk.Entry(self.time_frm, width=6)
        self.ent_end_time.insert(0, "23:59")
        self.ent_end_time.pack(side="left", padx=2)

        ttk.Label(self.time_frm, text="One-Shot:").pack(
            side="left", padx=(14, 4))
        self.cmb_oneshot_dur = ttk.Combobox(
            self.time_frm,
            values=[lbl for lbl, _ in ONE_SHOT_PRESETS],
            state="readonly", width=9,
        )
        self.cmb_oneshot_dur.set("1 hour")
        self.cmb_oneshot_dur.pack(side="left", padx=(0, 6))
        ttk.Button(
            self.time_frm, text="Start Now",
            command=self._start_one_shot, width=10,
        ).pack(side="left")

        # Retention row
        row_ret = ttk.Frame(frm)
        row_ret.pack(fill="x", pady=(4, 0))
        ttk.Label(row_ret, text="Keep logs:").pack(side="left")
        self.spn_retain = ttk.Spinbox(
            row_ret, from_=1, to=3650, width=6,
            command=self._on_retain_changed,
        )
        self.spn_retain.set(
            str(self.cfg.get("logging", {}).get("retain_days", 90)))
        self.spn_retain.pack(side="left", padx=4)
        ttk.Label(row_ret, text="days").pack(side="left")

        ttk.Button(
            row_ret, text="Purge Now",
            command=self._purge_now, width=10,
        ).pack(side="left", padx=(16, 0))

        self.lbl_purge_result = ttk.Label(row_ret, text="",
                                           foreground=_MUTED,
                                           font=("Segoe UI", 8))
        self.lbl_purge_result.pack(side="left", padx=8)

    # =========================================================================
    # Event handlers
    # =========================================================================

    def _on_start(self):
        self._paused = False
        self.btn_pause.config(text="⏸  Pause")
        self.cfg.setdefault("logging", {})["enabled"] = True
        self.logging_engine.enabled = True
        self._toast("Logging started")

    def _on_stop(self):
        self._paused = False
        self.btn_pause.config(text="⏸  Pause")
        self.cfg.setdefault("logging", {})["enabled"] = False
        self.logging_engine.stop_now()
        self._toast("Logging stopped", warn=True)

    def _on_pause(self):
        if not self._paused:
            self._paused = True
            self.btn_pause.config(text="▶  Resume")
            self.logging_engine.stop_now()
            self._toast("Logging paused", warn=True)
        else:
            self._paused = False
            self.btn_pause.config(text="⏸  Pause")
            self.logging_engine.enabled = True
            self._toast("Logging resumed")

    def _on_interval_changed(self, event=None):
        text = self.cmb_interval.get()
        for label, sec in INTERVALS:
            if label == text:
                self.cfg.setdefault("logging", {})["interval_sec"] = sec
                self.logging_engine.interval_sec = sec
                self._set_dirty(True)
                break

    def _on_mode_changed(self, event=None):
        text = self.cmb_mode.get()
        for label, mode in MODES:
            if label == text:
                self.cfg.setdefault("logging", {})["mode"] = mode
                self.logging_engine.mode = mode
                self._set_dirty(True)
                break

    def _on_keep_awake_changed(self):
        self.cfg.setdefault("logging", {})["keep_awake"] = self.keep_awake_var.get()
        self._set_dirty(True)

    def _on_schedule_changed(self):
        lcfg = self.cfg.setdefault("logging", {})
        schedule = lcfg.setdefault("schedule", {})
        schedule["mode"] = self.schedule_mode_var.get()
        schedule["start_datetime"] = (
            f"{self.ent_start_date.get()} {self.ent_start_time.get()}")
        schedule["end_datetime"] = (
            f"{self.ent_end_date.get()} {self.ent_end_time.get()}")
        self._set_dirty(True)

    def _on_apply_settings(self):
        snapshot_bus = getattr(self.logging_engine, "snapshot_bus", None)
        self.logging_engine.configure(
            self.cfg,
            meters=self.logging_engine.meters,
            snapshot_bus=snapshot_bus,
        )
        self._set_dirty(False)
        self._toast("Settings saved")

    def _on_retain_changed(self):
        try:
            val = max(1, int(self.spn_retain.get()))
        except Exception:
            val = 90
        self.cfg.setdefault("logging", {})["retain_days"] = val
        self.logging_engine.retain_days = val

    def _start_one_shot(self):
        dur_label = self.cmb_oneshot_dur.get()
        minutes = next(
            (m for lbl, m in ONE_SHOT_PRESETS if lbl == dur_label), 60)
        self.schedule_mode_var.set("one_shot")
        self._on_schedule_changed()
        self.logging_engine.start_one_shot(minutes)
        self.cfg.setdefault("logging", {})["enabled"] = True
        self._paused = False
        self.btn_pause.config(text="⏸  Pause")
        self._toast(f"One-shot started: {dur_label}")

    def _open_folder(self):
        path = self.logging_engine.current_file_path
        target = path if path else os.getcwd()
        folder = os.path.dirname(target) if os.path.isfile(target) else target
        if os.path.exists(folder):
            os.startfile(folder)

    def _copy_path(self):
        path = self.logging_engine.current_file_path or ""
        if path:
            self.clipboard_clear()
            self.clipboard_append(path)
            self._toast("Path copied")
        else:
            self._toast("No file active", warn=True)

    def _purge_now(self):
        try:
            retain = max(1, int(self.spn_retain.get()))
        except Exception:
            retain = 90
        folder = (
            (self.cfg.get("logging", {}).get("folder", "") or "").strip()
            or ""
        )
        if not folder:
            from utils.paths import logs_dir
            folder = logs_dir()
        deleted = LoggingEngine._purge_old_logs(folder, retain)
        self.lbl_purge_result.config(
            text=f"Removed {deleted} folder(s)",
            foreground="#34d399" if deleted else _MUTED,
        )

    # =========================================================================
    # Search / presets
    # =========================================================================

    def _on_search_changed(self):
        q = self._search_var.get().strip().lower()
        for source_id, row_data in self._source_rows.items():
            param_frame = row_data.get("param_frame")
            if param_frame is None:
                continue
            for widget in param_frame.winfo_children():
                if isinstance(widget, ttk.Checkbutton):
                    lbl_text = str(widget.cget("text")).lower()
                    if q == "" or q in lbl_text:
                        widget.grid()
                    else:
                        widget.grid_remove()

    def _select_all_visible(self):
        q = self._search_var.get().strip().lower()
        for source_id, row_data in self._source_rows.items():
            param_frame = row_data.get("param_frame")
            if param_frame is None:
                continue
            for widget in param_frame.winfo_children():
                if isinstance(widget, ttk.Checkbutton):
                    lbl_text = str(widget.cget("text")).lower()
                    if q == "" or q in lbl_text:
                        var = row_data.get("key_vars", {}).get(
                            widget.cget("text"), None)
                        if var:
                            var.set(True)
            self._sync_source_keys_from_vars(source_id)

    def _select_none_visible(self):
        q = self._search_var.get().strip().lower()
        for source_id, row_data in self._source_rows.items():
            param_frame = row_data.get("param_frame")
            if param_frame is None:
                continue
            for widget in param_frame.winfo_children():
                if isinstance(widget, ttk.Checkbutton):
                    lbl_text = str(widget.cget("text")).lower()
                    if q == "" or q in lbl_text:
                        var = row_data.get("key_vars", {}).get(
                            widget.cget("text"), None)
                        if var:
                            var.set(False)
            self._sync_source_keys_from_vars(source_id)

    def _apply_preset(self, name: str):
        keys = KEY_PRESETS.get(name, [])
        for source_id in list(self._source_rows.keys()):
            self._set_source_keys(source_id, keys)
            row_data = self._source_rows.get(source_id, {})
            key_vars = row_data.get("key_vars", {})
            for k, v in key_vars.items():
                v.set(k in keys)
        self._set_dirty(True)
        self._toast(f"Preset applied: {name}")

    # =========================================================================
    # Toast feedback
    # =========================================================================

    def _toast(self, msg: str, warn: bool = False):
        color = "#f0a030" if warn else "#34d399"
        self.lbl_toast.config(text=msg, foreground=color)
        if self._toast_job:
            try:
                self.after_cancel(self._toast_job)
            except Exception:
                pass
        self._toast_job = self.after(
            3000, lambda: self.lbl_toast.config(text=""))

    # =========================================================================
    # Dirty / save state
    # =========================================================================

    def _set_dirty(self, dirty: bool):
        self._dirty = dirty
        if dirty:
            self.btn_save.state(["!disabled"])
            self.lbl_dirty_dot.config(foreground="#f0a030")
        else:
            self.btn_save.state(["disabled"])
            self.lbl_dirty_dot.config(foreground="#1f1f23")

    # =========================================================================
    # Source management
    # =========================================================================

    def _load_from_cfg(self):
        lcfg = self.cfg.get("logging", {})

        interval_sec = lcfg.get("interval_sec", 10)
        for label, sec in INTERVALS:
            if sec == interval_sec:
                self.cmb_interval.set(label)
                break

        mode = lcfg.get("mode", "instant")
        for label, m in MODES:
            if m == mode:
                self.cmb_mode.set(label)
                break

        self.keep_awake_var.set(lcfg.get("keep_awake", False))

        schedule = lcfg.get("schedule", {})
        self.schedule_mode_var.set(schedule.get("mode", "always"))

        start_dt = schedule.get("start_datetime", "")
        if start_dt and " " in start_dt:
            parts = start_dt.split(" ")
            self.ent_start_date.delete(0, tk.END)
            self.ent_start_date.insert(0, parts[0])
            self.ent_start_time.delete(0, tk.END)
            self.ent_start_time.insert(0, parts[1] if len(parts) > 1 else "00:00")

        end_dt = schedule.get("end_datetime", "")
        if end_dt and " " in end_dt:
            parts = end_dt.split(" ")
            self.ent_end_date.delete(0, tk.END)
            self.ent_end_date.insert(0, parts[0])
            self.ent_end_time.delete(0, tk.END)
            self.ent_end_time.insert(0, parts[1] if len(parts) > 1 else "23:59")

        duration = schedule.get("duration_sec")
        if duration:
            label = self._find_duration_label(duration)
            if label:
                self.cmb_oneshot_dur.set(label)

        retain = lcfg.get("retain_days", 90)
        try:
            self.spn_retain.set(str(max(1, int(retain))))
        except Exception:
            pass

        self._ensure_meter_defaults()
        self._refresh_source_rows(force=True)
        self._set_dirty(False)

    def _refresh_source_rows_if_needed(self):
        total_keys = tuple(self._get_total_output_keys())
        meter_ids = tuple(sorted(
            int(m.meter_id)
            for m in self.logging_engine.meters
            if getattr(m, "meter_id", None) is not None
        ))
        if total_keys != self._last_total_keys or meter_ids != self._last_meter_ids:
            self._refresh_source_rows(total_keys, meter_ids)

    def _refresh_source_rows(self, total_keys=None, meter_ids=None, force=False):
        if not hasattr(self, "_source_container"):
            return
        computed_total_keys = tuple(
            total_keys or self._get_total_output_keys())
        computed_meter_ids = tuple(
            meter_ids or sorted(
                int(m.meter_id)
                for m in self.logging_engine.meters
                if getattr(m, "meter_id", None) is not None
            )
        )
        if (not force
                and computed_total_keys == self._last_total_keys
                and computed_meter_ids == self._last_meter_ids):
            return

        for child in self._source_container.winfo_children():
            child.destroy()
        self._source_rows.clear()

        total_cfg = self._get_total_cfg()
        self._add_source_row(
            source_id="TOTAL",
            display_name="TOTAL (aggregated)",
            enabled=total_cfg["enabled"],
            keys=total_cfg["keys"],
            options=self._get_total_options(),
        )

        ttk.Separator(
            self._source_container, orient="horizontal"
        ).pack(fill="x", pady=(4, 8))

        for meter in self.logging_engine.meters:
            meter_id = getattr(meter, "meter_id", None)
            if meter_id is None:
                continue
            cfg = self._get_meter_cfg(meter_id)
            label = getattr(meter, "name", f"Meter {meter_id}")
            self._add_source_row(
                source_id=f"M{meter_id}",
                display_name=label,
                enabled=cfg["enabled"],
                keys=cfg["keys"],
                options=ALL_LOG_KEYS,
            )

        self._last_total_keys = computed_total_keys
        self._last_meter_ids = computed_meter_ids

    def _add_source_row(self, source_id, display_name, enabled, keys, options):
        """Add a collapsible source row with inline parameter checkboxes."""
        outer = ttk.Frame(self._source_container)
        outer.pack(fill="x", pady=(0, 4))

        # Header row
        hdr = ttk.Frame(outer)
        hdr.pack(fill="x")

        enabled_var = tk.BooleanVar(value=bool(enabled))
        ttk.Checkbutton(
            hdr, text=display_name, variable=enabled_var,
            command=lambda sid=source_id, v=enabled_var:
                self._set_source_enabled(sid, v.get()),
        ).pack(side="left")

        # Summary label
        summary_var = tk.StringVar(value=self._format_key_summary(keys))
        summary_lbl = ttk.Label(
            hdr, textvariable=summary_var,
            foreground=_MUTED, font=("Segoe UI", 8),
        )
        summary_lbl.pack(side="left", padx=(8, 0), fill="x", expand=True)

        # Toggle button
        expanded_var = tk.BooleanVar(value=False)
        toggle_btn = ttk.Button(
            hdr, text="▼ Parameters",
            command=lambda: _toggle(),
            width=13,
        )
        toggle_btn.pack(side="right", padx=(4, 0))

        # Parameter frame (hidden by default)
        param_frame = ttk.Frame(outer, padding=(24, 4, 4, 4))

        # Build key_vars dict (keyed by display label for search compat)
        key_vars: dict[str, tk.BooleanVar] = {}
        current_set = set(keys)
        cols = 4
        for idx, key in enumerate(options):
            var = tk.BooleanVar(value=key in current_set)
            lbl_text = key_label(key) or key
            cb = ttk.Checkbutton(
                param_frame, text=lbl_text, variable=var,
                command=lambda sid=source_id: self._sync_source_keys_from_vars(sid),
            )
            cb.grid(row=idx // cols, column=idx % cols,
                    sticky="w", padx=4, pady=1)
            key_vars[lbl_text] = var

        def _toggle():
            if expanded_var.get():
                expanded_var.set(False)
                param_frame.pack_forget()
                toggle_btn.config(text="▼ Parameters")
            else:
                expanded_var.set(True)
                param_frame.pack(fill="x")
                toggle_btn.config(text="▲ Parameters")

        self._source_rows[source_id] = {
            "enabled_var": enabled_var,
            "summary_var": summary_var,
            "key_vars": key_vars,
            "param_frame": param_frame,
            "options": options,
        }

    def _sync_source_keys_from_vars(self, source_id: str):
        row_data = self._source_rows.get(source_id)
        if row_data is None:
            return
        options = row_data.get("options", [])
        key_vars = row_data.get("key_vars", {})
        # Map display label → canonical key
        label_to_key = {key_label(k) or k: k for k in options}
        selected = [
            label_to_key[lbl]
            for lbl, var in key_vars.items()
            if var.get() and lbl in label_to_key
        ]
        self._set_source_keys(source_id, selected)
        # Update summary
        summary_var = row_data.get("summary_var")
        if summary_var:
            summary_var.set(self._format_key_summary(selected or options[:3]))

    # -------------------------------------------------------------------------
    # Config accessors (unchanged from original)
    # -------------------------------------------------------------------------

    def _get_total_cfg(self):
        sources = self.cfg.setdefault("logging", {}).setdefault("sources", {})
        total_cfg = sources.get("total", {})
        if isinstance(total_cfg, bool):
            total_cfg = {"enabled": total_cfg}
        total_cfg.setdefault("enabled", True)
        keys = self._canonicalize_key_list(total_cfg.get("keys", []))
        if not keys:
            keys = self._get_total_output_keys()
        total_cfg["keys"] = keys
        sources["total"] = total_cfg
        return total_cfg

    def _ensure_meter_defaults(self):
        sources = self.cfg.setdefault("logging", {}).setdefault("sources", {})
        meters_cfg = sources.setdefault("meters", {})
        default_keys = meters_cfg.get("default_keys")
        if not default_keys:
            default_keys = self.cfg.get("logging", {}).get(
                "keys", DEFAULT_LOG_KEYS)
        meters_cfg["default_keys"] = (
            self._canonicalize_key_list(default_keys) or DEFAULT_LOG_KEYS)
        meters_cfg.setdefault("per_meter", {})

    def _get_meter_cfg(self, meter_id):
        sources = self.cfg.setdefault("logging", {}).setdefault("sources", {})
        meters_cfg = sources.setdefault("meters", {})
        self._ensure_meter_defaults()
        per_meter = meters_cfg.setdefault("per_meter", {})
        key = str(meter_id)
        entry = per_meter.get(key, {})
        entry.setdefault("enabled", True)
        entry["keys"] = (
            self._canonicalize_key_list(entry.get("keys", []))
            or meters_cfg.get("default_keys", DEFAULT_LOG_KEYS)
        )
        per_meter[key] = entry
        return entry

    def _get_meter_defaults(self):
        sources = self.cfg.setdefault("logging", {}).setdefault("sources", {})
        meters_cfg = sources.setdefault("meters", {})
        self._ensure_meter_defaults()
        return meters_cfg.get("default_keys", DEFAULT_LOG_KEYS)

    def _canonicalize_key_list(self, keys):
        normalized: list[str] = []
        for key in keys or []:
            canon = canonicalize_log_key(key)
            if canon and canon not in normalized:
                normalized.append(canon)
        return normalized

    def _get_total_options(self):
        options = self._get_total_output_keys()
        return options or DEFAULT_LOG_KEYS

    def _get_source_keys(self, source_id):
        if source_id == "TOTAL":
            return self._get_total_cfg()["keys"]
        try:
            meter_id = int(source_id[1:])
        except Exception:
            return []
        return self._get_meter_cfg(meter_id)["keys"]

    def _set_source_keys(self, source_id, keys):
        canonical = self._canonicalize_key_list(keys)
        if source_id == "TOTAL":
            cfg = self._get_total_cfg()
            cfg["keys"] = canonical or self._get_total_output_keys()
        else:
            meter_id = int(source_id[1:])
            cfg = self._get_meter_cfg(meter_id)
            fallback = cfg.get("keys") or list(self._get_meter_defaults())
            cfg["keys"] = canonical or fallback
        self._set_dirty(True)

    def _set_source_enabled(self, source_id, enabled):
        if source_id == "TOTAL":
            cfg = self._get_total_cfg()
            cfg["enabled"] = bool(enabled)
        else:
            meter_id = int(source_id[1:])
            cfg = self._get_meter_cfg(meter_id)
            cfg["enabled"] = bool(enabled)
        self._set_dirty(True)

    def _format_key_summary(self, keys):
        cleaned = [k for k in keys or [] if k]
        if not cleaned:
            return "No parameters selected"
        labels = [key_label(k) or k for k in cleaned[:4]]
        if len(cleaned) > 4:
            return f"{', '.join(labels)}  +{len(cleaned) - 4} more"
        return ", ".join(labels)

    def _find_duration_label(self, duration_sec):
        for label, minutes in ONE_SHOT_PRESETS:
            if minutes * 60 == duration_sec:
                return label
        return None

    def _get_total_output_keys(self):
        slots = (self.cfg.get("total_custom") or {}).get("slots", []) or []
        seen: list[str] = []
        for slot in slots[:16]:
            key = canonical_key(slot.get("output_key"))
            if key and key not in seen:
                seen.append(key)
        return seen

    # =========================================================================
    # Recent records preview
    # =========================================================================

    def _refresh_preview(self):
        """Read last 5 rows from the current log file / historian."""
        engine = self.logging_engine
        historian = getattr(engine, "_historian", None)

        if historian is not None:
            self._load_preview_from_historian(historian)
        elif engine.current_file_path and os.path.isfile(engine.current_file_path):
            self._load_preview_from_csv(engine.current_file_path)
        else:
            self._set_preview_empty("No active log file")

    def _load_preview_from_csv(self, path: str):
        try:
            with open(path, newline="", encoding="utf-8-sig") as f:
                reader = csv.reader(f)
                all_rows = [r for r in reader
                            if r and not str(r[0]).startswith("#")]
            if len(all_rows) < 2:
                self._set_preview_empty("File has no data rows yet")
                return
            headers = all_rows[0]
            data = all_rows[max(1, len(all_rows) - 5):]
            self._populate_preview(headers, data)
            self.lbl_preview_info.config(
                text=f"Last {len(data)} rows — {os.path.basename(path)}")
        except Exception as exc:
            self._set_preview_empty(f"Read error: {exc}")

    def _load_preview_from_historian(self, historian):
        try:
            to_ts = datetime.now().timestamp()
            from_ts = to_ts - 3600  # last hour
            rows = historian.query_range(from_ts, to_ts)
            if not rows:
                self._set_preview_empty("No historian data in last hour")
                return
            recent = rows[-5:]
            if not recent:
                self._set_preview_empty("No recent rows")
                return
            headers = list(recent[0].keys())
            data = [[str(r.get(h, "")) for h in headers] for r in recent]
            self._populate_preview(headers, data)
            self.lbl_preview_info.config(
                text=f"Last {len(data)} rows — historian")
        except Exception as exc:
            self._set_preview_empty(f"Historian error: {exc}")

    def _populate_preview(self, headers: list[str], data: list[list[str]]):
        tree = self._preview_tree
        tree.delete(*tree.get_children())
        cols = tuple(headers)
        tree["columns"] = cols
        for col in cols:
            tree.heading(col, text=col)
            tree.column(col, width=max(70, len(col) * 9), minwidth=50)
        for row in data:
            tree.insert("", "end", values=row)
        self._preview_cols = cols

    def _set_preview_empty(self, msg: str):
        tree = self._preview_tree
        tree.delete(*tree.get_children())
        tree["columns"] = ("info",)
        tree.heading("info", text="")
        tree.column("info", width=300)
        tree.insert("", "end", values=(msg,))
        self.lbl_preview_info.config(text="")

    # =========================================================================
    # View update (called every ~500 ms by main_window poll)
    # =========================================================================

    def update_view(self):
        engine = self.logging_engine
        state = engine.state

        # ── State chip ───────────────────────────────────────────────────────
        chip_key = state if state in _CHIP else "DISABLED"
        if self._paused:
            chip_key = "PAUSED"
        bg, fg = _CHIP[chip_key]
        self.chip_state.config(
            text=_chip_text(chip_key), bg=bg, fg=fg)

        # ── Disk chip (shown only when warning) ──────────────────────────────
        try:
            db_mb = getattr(engine, "db_size_mb", 0.0)
            folder_mb = engine.folder_size_mb
            warn_mb = max(db_mb, folder_mb)
        except Exception:
            warn_mb = 0.0

        if warn_mb > 200:
            label = (
                "⚠ >500 MB" if warn_mb > 500
                else f"⚠ {warn_mb:.0f} MB"
            )
            self.chip_disk.config(text=label)
            if not self.chip_disk.winfo_ismapped():
                self.chip_disk.pack(side="left", padx=(0, 4))
        else:
            if self.chip_disk.winfo_ismapped():
                self.chip_disk.pack_forget()

        # ── Error chip ───────────────────────────────────────────────────────
        err = engine.last_error
        if err:
            self.chip_err.config(text="✕ ERROR")
            if not self.chip_err.winfo_ismapped():
                self.chip_err.pack(side="left", padx=(0, 4))
        else:
            if self.chip_err.winfo_ismapped():
                self.chip_err.pack_forget()

        # ── Metrics ──────────────────────────────────────────────────────────
        if engine.next_write_time and state == "RUNNING":
            delta = (engine.next_write_time - datetime.now()).total_seconds()
            self._v_next.set(f"{max(0.0, delta):.1f}s")
        else:
            self._v_next.set("—")

        self._v_last.set(_fmt_age(getattr(engine, "last_write_time", None)))
        self._v_rows.set(f"{engine.rows_written_today:,}")

        errs = getattr(engine, "write_error_count", getattr(engine, "_write_error_count", 0))
        self._v_errs.set(str(errs) if errs else "0")

        dropped = getattr(engine, "dropped_write_count", getattr(engine, "_queue_full_count", 0))
        clock_jumps = getattr(engine, "clock_jump_count", getattr(engine, "_clock_jump_count", 0))
        self._v_skipped.set(f"{dropped} / clk {clock_jumps}" if clock_jumps else str(dropped))

        db_mb = getattr(engine, "db_size_mb", 0.0)
        db_rows = getattr(engine, "db_row_count", 0)
        if db_mb and db_mb > 0.0:
            self._v_storage.set(f"DB {db_mb:.1f} MB  ({db_rows:,} rows)")
        else:
            try:
                self._v_storage.set(f"CSV {engine.folder_size_mb:.1f} MB")
            except Exception:
                self._v_storage.set("—")

        # ── File path ────────────────────────────────────────────────────────
        fp = engine.current_file_path or "No file"
        # Truncate from left if long
        if len(fp) > 70:
            fp = "…" + fp[-67:]
        self.lbl_file.config(text=fp)

        # ── Error label (below chip row if needed) ───────────────────────────
        # (errors surfaced via chip only — no separate label)

        self._refresh_source_rows_if_needed()

    # =========================================================================
    # Excel export (preserved exactly from original)
    # =========================================================================

    def _export_excel(self):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            messagebox.showerror(
                "Missing package",
                "openpyxl is not installed.\nRun: pip install openpyxl",
            )
            return

        historian = getattr(self.logging_engine, "_historian", None)

        win = tk.Toplevel(self)
        win.title("Export Logs to Excel")
        win.resizable(False, False)
        win.transient(self)
        win.grab_set()

        frm = ttk.Frame(win, padding=12)
        frm.pack(fill="both", expand=True)

        backend_lbl = "SQLite historian" if historian else "CSV files"
        ttk.Label(frm, text=f"Data source: {backend_lbl}",
                  foreground="#94a3b8",
                  font=("Segoe UI", 8)).grid(
            row=0, column=0, columnspan=2, sticky="w", pady=(0, 6))

        ttk.Label(frm, text="From date (YYYY-MM-DD):").grid(
            row=1, column=0, sticky="w", pady=4)
        ent_from = ttk.Entry(frm, width=14)
        ent_from.insert(
            0, (datetime.now() - timedelta(days=6)).strftime("%Y-%m-%d"))
        ent_from.grid(row=1, column=1, padx=8, sticky="w")

        ttk.Label(frm, text="To date (YYYY-MM-DD):").grid(
            row=2, column=0, sticky="w", pady=4)
        ent_to = ttk.Entry(frm, width=14)
        ent_to.insert(0, datetime.now().strftime("%Y-%m-%d"))
        ent_to.grid(row=2, column=1, padx=8, sticky="w")

        lbl_status = ttk.Label(frm, text="", foreground="gray")
        lbl_status.grid(row=3, column=0, columnspan=2, pady=4, sticky="w")

        def _do_export():
            from utils.paths import logs_dir as _logs_dir
            import re as _re

            try:
                d_from = datetime.strptime(ent_from.get().strip(), "%Y-%m-%d")
                d_to   = datetime.strptime(ent_to.get().strip(),   "%Y-%m-%d")
            except ValueError:
                lbl_status.config(text="Invalid date format.", foreground="red")
                return
            if d_from.date() > d_to.date():
                lbl_status.config(
                    text="'From' must be <= 'To'.", foreground="red")
                return

            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            sheets_created = 0

            hdr_font  = Font(bold=True, color="FFFFFF")
            hdr_fill  = PatternFill("solid", fgColor="1e3a5f")
            hdr_align = Alignment(horizontal="center")

            def _style_header(ws_obj, headers):
                for col_idx, h in enumerate(headers, start=1):
                    cell = ws_obj.cell(row=1, column=col_idx, value=h)
                    cell.font      = hdr_font
                    cell.fill      = hdr_fill
                    cell.alignment = hdr_align

            if historian is not None:
                try:
                    from_epoch = d_from.replace(
                        hour=0, minute=0, second=0).timestamp()
                    to_epoch   = d_to.replace(
                        hour=23, minute=59, second=59).timestamp()
                    lbl_status.config(
                        text="Querying historian…", foreground="gray")
                    win.update_idletasks()
                    rows = historian.query_range(from_epoch, to_epoch)
                    if not rows:
                        lbl_status.config(
                            text="No data found for that date range.",
                            foreground="orange")
                        return
                    from collections import defaultdict
                    groups: dict = defaultdict(list)
                    for r in rows:
                        day = r["ts"][:10]
                        groups[(day, r["source"])].append(r)
                    for (day, source), grp_rows in sorted(groups.items()):
                        headers, out_rows = historian.rows_to_csv_lines(
                            grp_rows, source)
                        sname = f"{source}_{day}"[:31]
                        ws = wb.create_sheet(title=sname)
                        _style_header(ws, headers)
                        for r in out_rows:
                            ws.append([r.get(h, "") for h in headers])
                        sheets_created += 1
                except Exception as exc:
                    lbl_status.config(
                        text=f"Historian query failed: {exc}",
                        foreground="red")
                    return
            else:
                folder = (
                    (self.cfg.get("logging", {}).get("folder", "") or "").strip()
                    or _logs_dir()
                )
                day_pattern = _re.compile(r"^\d{4}-\d{2}-\d{2}$")
                try:
                    day_dirs = sorted(
                        n for n in os.listdir(folder)
                        if day_pattern.match(n)
                        and os.path.isdir(os.path.join(folder, n))
                    )
                except Exception:
                    day_dirs = []

                for day_name in day_dirs:
                    try:
                        day_date = datetime.strptime(
                            day_name, "%Y-%m-%d").date()
                    except ValueError:
                        continue
                    if not (d_from.date() <= day_date <= d_to.date()):
                        continue
                    day_path = os.path.join(folder, day_name)
                    try:
                        csv_files = sorted(
                            f for f in os.listdir(day_path)
                            if f.lower().endswith(".csv")
                        )
                    except Exception:
                        continue
                    for csv_name in csv_files:
                        csv_path = os.path.join(day_path, csv_name)
                        sname = os.path.splitext(csv_name)[0][:31]
                        ws = wb.create_sheet(title=sname)
                        try:
                            with open(csv_path, newline="",
                                      encoding="utf-8-sig") as f:
                                reader = csv.reader(f)
                                rows_raw = list(reader)
                            data_rows = [
                                r for r in rows_raw
                                if not (r and str(r[0]).startswith("#"))
                            ]
                            if data_rows:
                                _style_header(ws, data_rows[0])
                                for r in data_rows[1:]:
                                    ws.append(r)
                            sheets_created += 1
                        except Exception:
                            pass

            if sheets_created == 0:
                lbl_status.config(
                    text="No log data found in that date range.",
                    foreground="orange")
                return

            save_path = filedialog.asksaveasfilename(
                parent=win,
                title="Save Excel Export",
                defaultextension=".xlsx",
                filetypes=[("Excel Workbook", "*.xlsx")],
                initialfile=(
                    f"logs_{d_from.strftime('%Y-%m-%d')}"
                    f"_{d_to.strftime('%Y-%m-%d')}.xlsx"
                ),
            )
            if not save_path:
                lbl_status.config(text="Export cancelled.", foreground="gray")
                return

            try:
                wb.save(save_path)
                lbl_status.config(
                    text=f"Exported {sheets_created} sheet(s) successfully.",
                    foreground="green")
                win.after(1800, win.destroy)
            except Exception as exc:
                lbl_status.config(
                    text=f"Save failed: {exc}", foreground="red")

        btn_row = ttk.Frame(win, padding=(12, 4))
        btn_row.pack(fill="x")
        ttk.Button(btn_row, text="Cancel",
                   command=win.destroy).pack(side="right", padx=(8, 12))
        ttk.Button(btn_row, text="Export",
                   command=_do_export).pack(side="right")
